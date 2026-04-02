from __future__ import annotations

import csv
import json
import shutil
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .memory.writeback import bootstrap_memory_files, load_machine_state
from .project import ProjectPaths, find_repo_root


FEED_VERSION = "excel_console_feed_v5_no_scripts"
WORKBOOK_NAME = "ResearchConsole.xlsx"
WORKSHEET_NAMES = ("Control", "Overview", "Strategies", "Experiments", "Runs", "Artifacts")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=18, bold=True)
_BUTTON_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
_BUTTON_FONT = Font(color="0B5394", bold=True, underline="single")
_CAUTION_FILL = PatternFill(fill_type="solid", fgColor="FCE5CD")
_SUCCESS_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
_DANGER_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
_NEUTRAL_FILL = PatternFill(fill_type="solid", fgColor="EDEDED")
_CARD_FILL = PatternFill(fill_type="solid", fgColor="F7F9FC")
_SUBTLE_FILL = PatternFill(fill_type="solid", fgColor="F3F6FA")
_THIN_SIDE = Side(style="thin", color="D9D9D9")
_BOX_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_SECTION_TITLE_FONT = Font(size=12, bold=True, color="FFFFFF")
_LABEL_FONT = Font(bold=True, color="2F2F2F")
_SMALL_MUTED_FONT = Font(size=10, color="666666")
_METRIC_VALUE_FONT = Font(size=16, bold=True, color="1F1F1F")

_STRATEGY_LABELS = {
    "f1_elasticnet_v1": "F1 主线",
    "f2_structured_latent_factor_v1": "F2 挑战者",
    "baseline_limit_up": "基线对照",
    "r1_predictive_error_overlay_v1": "R1.1 风控",
    "r1_predictive_error_overlay_v2": "R1.2 风控",
    "risk_constrained_limit_up": "风控旧分支",
    "tighter_entry_limit_up": "收紧入场旧分支",
    "hybrid_f1_5_frozen_sidecar": "Hybrid F1.5 Frozen Sidecar",
    "legacy_single_branch": "旧兼容分支",
}


@dataclass(frozen=True)
class ExcelConsolePaths:
    root: Path
    feed_dir: Path
    workbook_path: Path
    manifest_path: Path
    notes_path: Path


def _utc_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat()


def _read_json(path: Path, default: dict[str, Any] | None = None) -> dict[str, Any]:
    if not path.exists():
        return dict(default or {})
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return dict(default or {})


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        stripped = raw_line.strip()
        if not stripped:
            continue
        try:
            payload = json.loads(stripped)
        except json.JSONDecodeError:
            continue
        if isinstance(payload, dict):
            rows.append(payload)
    return rows


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _first(iterable: Iterable[str]) -> str:
    for item in iterable:
        text = _normalize_text(item)
        if text:
            return text
    return ""


def _display_strategy_name(strategy_id: str, raw_name: Any) -> str:
    strategy_text = _normalize_text(strategy_id)
    name_text = _normalize_text(raw_name)
    if strategy_text in _STRATEGY_LABELS:
        return _STRATEGY_LABELS[strategy_text]
    if "?" in name_text or "\ufffd" in name_text:
        return strategy_text or name_text
    return name_text or strategy_text


def _read_report(path_text: str) -> dict[str, Any]:
    if not path_text:
        return {}
    try:
        return _read_json(Path(path_text), default={})
    except OSError:
        return {}


def _coerce_float(value: Any) -> float | None:
    try:
        if value in ("", None):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _fmt_pct(value: Any) -> str:
    number = _coerce_float(value)
    if number is None:
        return ""
    return f"{number * 100:.2f}%"


def _fmt_float(value: Any, digits: int = 2) -> str:
    number = _coerce_float(value)
    if number is None:
        return ""
    return f"{number:.{digits}f}"


def _card_fill_for_status(status: str) -> PatternFill:
    text = _normalize_text(status).lower()
    if any(token in text for token in ("pass", "promote", "ready", "mainline", "accepted")):
        return _SUCCESS_FILL
    if any(token in text for token in ("reject", "blocked", "above", "fail")):
        return _DANGER_FILL
    if any(token in text for token in ("mixed", "hold", "continue", "challenger")):
        return _CAUTION_FILL
    return _CARD_FILL


def _excel_console_paths(paths: ProjectPaths) -> ExcelConsolePaths:
    root = paths.artifacts_dir / "excel"
    feed_dir = root / "feed"
    return ExcelConsolePaths(
        root=root,
        feed_dir=feed_dir,
        workbook_path=root / WORKBOOK_NAME,
        manifest_path=feed_dir / "manifest.json",
        notes_path=root / "EXCEL_CONSOLE_NOTES.md",
    )


def _ensure_excel_dirs(console_paths: ExcelConsolePaths) -> None:
    console_paths.root.mkdir(parents=True, exist_ok=True)
    console_paths.feed_dir.mkdir(parents=True, exist_ok=True)


def _cleanup_legacy_excel_artifacts(console_paths: ExcelConsolePaths) -> list[str]:
    removed: list[str] = []
    legacy_patterns = (
        "*.cmd",
        "*.bat",
        "*.ps1",
        "*.vbs",
        "*.wsf",
        "ResearchConsole*.xlsm",
    )
    for pattern in legacy_patterns:
        for path in console_paths.root.glob(pattern):
            if path.is_file():
                path.unlink(missing_ok=True)
                removed.append(str(path))
    for legacy_dir_name in ("actions", "_staging"):
        legacy_dir = console_paths.root / legacy_dir_name
        if legacy_dir.exists():
            shutil.rmtree(legacy_dir, ignore_errors=True)
            removed.append(str(legacy_dir))
    return removed


def _write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            normalized: dict[str, Any] = {}
            for column in columns:
                value = row.get(column, "")
                if isinstance(value, str):
                    normalized[column] = _clean_export_text(value)
                else:
                    normalized[column] = value
            writer.writerow(normalized)
    return path


def _write_text(path: Path, text: str) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")
    return path


def _clean_export_text(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\u0000", "")
    return text.strip()


def _current_mainline_id(state: dict[str, Any]) -> str:
    primary = [str(item).strip() for item in (state.get("current_primary_strategy_ids", []) or []) if str(item).strip()]
    return primary[0] if primary else "f1_elasticnet_v1"


def _current_challenger_id(state: dict[str, Any]) -> str:
    preferred = [
        str(state.get("next_build_target", "")).strip(),
        "f2_structured_latent_factor_v1",
    ]
    secondary = [str(item).strip() for item in (state.get("current_secondary_strategy_ids", []) or []) if str(item).strip()]
    for candidate in [*preferred, *secondary]:
        if candidate and candidate != _current_mainline_id(state):
            return candidate
    return "f2_structured_latent_factor_v1"


def _strategy_role(state: dict[str, Any], strategy_id: str, track: str = "") -> str:
    strategy_text = _normalize_text(strategy_id)
    if not strategy_text:
        return "other"
    if strategy_text == _current_mainline_id(state):
        return "mainline"
    if strategy_text == "baseline_limit_up" or _normalize_text(track) == "control":
        return "control"
    if strategy_text in {str(item).strip() for item in (state.get("current_blocked_strategy_ids", []) or []) if str(item).strip()}:
        return "blocked"
    if strategy_text in {str(item).strip() for item in (state.get("current_rejected_strategy_ids", []) or []) if str(item).strip()}:
        return "rejected"
    if strategy_text == _current_challenger_id(state):
        return "challenger"
    if _normalize_text(track) in {"candidate", "secondary"}:
        return "challenger"
    return "other"


def _role_label(role: str) -> str:
    labels = {
        "mainline": "主线",
        "challenger": "挑战者",
        "control": "对照",
        "blocked": "阻塞",
        "rejected": "已否决",
        "other": "其它",
    }
    return labels.get(role, role or "其它")


def _role_fill(role: str) -> PatternFill:
    mapping = {
        "mainline": _SUCCESS_FILL,
        "challenger": _CAUTION_FILL,
        "control": _NEUTRAL_FILL,
        "blocked": _DANGER_FILL,
        "rejected": _DANGER_FILL,
        "other": _CARD_FILL,
    }
    return mapping.get(role, _CARD_FILL)


def _canonical_blocker_text(state: dict[str, Any]) -> str:
    blocker = _normalize_text(state.get("current_blocker"))
    if "?" in blocker or "\ufffd" in blocker:
        f2_decision = _normalize_text(state.get("f2_verify_decision"))
        if f2_decision == "keep_f2_challenger":
            return "F2.1 shared-shell Top6 max_drawdown 33.58% remains above 30.00%."
    return blocker or "当前没有记录 blocker。"


def _latest_conclusion_text(state: dict[str, Any]) -> str:
    explicit = _normalize_text(state.get("current_conclusion"))
    if explicit and "?" not in explicit and "\ufffd" not in explicit:
        return explicit
    decision = _normalize_text(state.get("f2_verify_decision"))
    if decision == "keep_f2_challenger":
        return "F2.1 已跑通，但当前仍只是挑战者，尚未取代 F1 主线。"
    if decision == "promote_f2_next":
        return "F2.1 已升为下一条主推挑战者。"
    return "当前还没有新的主线替代结论。"


def _next_action_text(state: dict[str, Any]) -> str:
    explicit = _normalize_text(state.get("next_priority_action"))
    if explicit and "?" not in explicit and "\ufffd" not in explicit:
        return explicit
    return "Run one more bounded F2 variant before widening the search."


def _research_scope_text(state: dict[str, Any]) -> str:
    explicit = _normalize_text(state.get("current_validation_scope"))
    if explicit and "?" not in explicit and "\ufffd" not in explicit:
        return explicit
    return "同一核心股票池 + 同一 shared shell + Top6。"


def _strategy_lookup_by_id(rows: list[dict[str, Any]], strategy_id: str) -> dict[str, Any]:
    return next((row for row in rows if _normalize_text(row.get("strategy_id")) == strategy_id), {})


def _overview_rows(state: dict[str, Any], *, workbook_mode: str, workbook_path: Path) -> list[dict[str, Any]]:
    return [
        {"section": "研究状态", "key": "当前阶段", "value": _normalize_text(state.get("current_phase")), "display_order": 10},
        {"section": "研究状态", "key": "当前主线", "value": _display_strategy_name(_current_mainline_id(state), ""), "display_order": 20},
        {"section": "研究状态", "key": "当前挑战者", "value": _display_strategy_name(_current_challenger_id(state), ""), "display_order": 30},
        {"section": "研究状态", "key": "当前 blocker", "value": _canonical_blocker_text(state), "display_order": 40},
        {"section": "研究状态", "key": "最新结论", "value": _latest_conclusion_text(state), "display_order": 50},
        {"section": "研究状态", "key": "下一步", "value": _next_action_text(state), "display_order": 60},
        {"section": "研究状态", "key": "验证口径", "value": _research_scope_text(state), "display_order": 70},
        {
            "section": "平台状态",
            "key": "最新验证能力",
            "value": _normalize_text(state.get("last_verified_capability")),
            "display_order": 80,
        },
        {
            "section": "平台状态",
            "key": "核心股票池",
            "value": _normalize_text(state.get("canonical_universe_id")),
            "display_order": 90,
        },
        {
            "section": "平台状态",
            "key": "控制台模式",
            "value": workbook_mode,
            "display_order": 100,
        },
        {
            "section": "平台状态",
            "key": "工作簿路径",
            "value": str(workbook_path),
            "display_order": 110,
        },
    ]


def _research_summary_rows(state: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {"section": "研究状态", "key": "当前阶段", "value": _normalize_text(state.get("current_phase")), "display_order": 10},
        {"section": "研究状态", "key": "当前主线", "value": _display_strategy_name(_current_mainline_id(state), ""), "display_order": 20},
        {"section": "研究状态", "key": "当前挑战者", "value": _display_strategy_name(_current_challenger_id(state), ""), "display_order": 30},
        {"section": "研究状态", "key": "当前 blocker", "value": _canonical_blocker_text(state), "display_order": 40},
        {"section": "研究状态", "key": "最新结论", "value": _latest_conclusion_text(state), "display_order": 50},
        {"section": "研究状态", "key": "下一步", "value": _next_action_text(state), "display_order": 60},
        {"section": "研究状态", "key": "验证口径", "value": _research_scope_text(state), "display_order": 70},
    ]


def _control_cards_rows(state: dict[str, Any], strategy_metrics_rows: list[dict[str, Any]], *, exported_at: str) -> list[dict[str, Any]]:
    primary = _strategy_lookup_by_id(strategy_metrics_rows, _current_mainline_id(state))
    challenger = _strategy_lookup_by_id(strategy_metrics_rows, _current_challenger_id(state))
    control = _strategy_lookup_by_id(strategy_metrics_rows, "baseline_limit_up")
    return [
        {
            "slot": "card_mainline",
            "title": "当前主线",
            "value": _display_strategy_name(_current_mainline_id(state), primary.get("name")),
            "status": "mainline",
            "display_order": 10,
        },
        {
            "slot": "card_blocker",
            "title": "当前 blocker",
            "value": _canonical_blocker_text(state),
            "status": "blocked",
            "display_order": 20,
        },
        {
            "slot": "card_next_step",
            "title": "下一步",
            "value": _next_action_text(state),
            "status": "next_step",
            "display_order": 30,
        },
        {
            "slot": "card_conclusion",
            "title": "当前结论",
            "value": _latest_conclusion_text(state),
            "status": "conclusion",
            "display_order": 40,
        },
        {
            "slot": "metric_f1",
            "title": "F1 核心指标",
            "value": f"{_fmt_pct(primary.get('annualized_return')) or 'n/a'} / {_fmt_pct(primary.get('max_drawdown')) or 'n/a'}",
            "status": "mainline",
            "display_order": 50,
        },
        {
            "slot": "metric_f2",
            "title": "F2 核心指标",
            "value": f"{_fmt_pct(challenger.get('annualized_return')) or 'n/a'} / {_fmt_pct(challenger.get('max_drawdown')) or 'n/a'}",
            "status": "challenger",
            "display_order": 60,
        },
        {
            "slot": "metric_control",
            "title": "对照指标",
            "value": f"{_fmt_pct(control.get('annualized_return')) or 'n/a'} / {_fmt_pct(control.get('max_drawdown')) or 'n/a'}",
            "status": "control",
            "display_order": 70,
        },
        {
            "slot": "metric_export",
            "title": "最近导出时间",
            "value": exported_at,
            "status": "info",
            "display_order": 80,
        },
    ]


def _strategy_metrics_rows(state: dict[str, Any]) -> list[dict[str, Any]]:
    candidates = {
        _normalize_text(item.get("strategy_id")): item
        for item in (state.get("strategy_candidates", []) or [])
        if isinstance(item, dict) and _normalize_text(item.get("strategy_id"))
    }
    rows: list[dict[str, Any]] = []

    def add_row(
        *,
        strategy_id: str,
        metrics: dict[str, Any],
        decision: str = "",
        classification: str = "",
        artifact_ref: str = "",
        track: str = "",
    ) -> None:
        candidate = candidates.get(strategy_id, {})
        rows.append(
            {
                "strategy_id": strategy_id,
                "name": _display_strategy_name(strategy_id, candidate.get("name")),
                "role": _strategy_role(state, strategy_id, track or _normalize_text(candidate.get("track"))),
                "track": track or _normalize_text(candidate.get("track")),
                "decision": decision or _normalize_text(candidate.get("decision")),
                "classification": classification,
                "annualized_return": _coerce_float(metrics.get("annualized_return")),
                "max_drawdown": _coerce_float(metrics.get("max_drawdown")),
                "sharpe_ratio": _coerce_float(metrics.get("sharpe_ratio")),
                "calmar_ratio": _coerce_float(metrics.get("calmar_ratio")),
                "turnover": _coerce_float(metrics.get("turnover")),
                "win_rate": _coerce_float(metrics.get("win_rate")),
                "artifact_ref": artifact_ref,
            }
        )

    f1_report = _read_report(_normalize_text(state.get("f1_verifier_report_path")))
    if f1_report:
        add_row(
            strategy_id="baseline_limit_up",
            metrics=dict(f1_report.get("control_metrics") or {}),
            decision="control",
            classification=_normalize_text(f1_report.get("classification")),
            artifact_ref=_normalize_text(f1_report.get("artifact_paths", {}).get("plot_path")),
            track="control",
        )
        add_row(
            strategy_id="f1_elasticnet_v1",
            metrics=dict(f1_report.get("f1_metrics") or {}),
            decision=_normalize_text(f1_report.get("decision")),
            classification=_normalize_text(f1_report.get("classification")),
            artifact_ref=_normalize_text(f1_report.get("artifact_paths", {}).get("plot_path")),
            track="primary",
        )

    r1_report = _read_report(_normalize_text(state.get("r1_verify_report_path")))
    if r1_report:
        regime_profile = _normalize_text(r1_report.get("regime_control", {}).get("profile"))
        strategy_id = "r1_predictive_error_overlay_v2" if regime_profile.endswith("_v2") else "r1_predictive_error_overlay_v1"
        add_row(
            strategy_id=strategy_id,
            metrics=dict(r1_report.get("r1_metrics") or {}),
            decision=_normalize_text(r1_report.get("decision")),
            classification=_normalize_text(r1_report.get("classification")),
            artifact_ref=_normalize_text(r1_report.get("artifact_paths", {}).get("plot_path")),
            track="candidate",
        )

    f2_report = _read_report(_normalize_text(state.get("f2_verify_report_path")))
    if f2_report:
        add_row(
            strategy_id="f2_structured_latent_factor_v1",
            metrics=dict(f2_report.get("f2_metrics") or {}),
            decision=_normalize_text(f2_report.get("decision")),
            classification=_normalize_text(f2_report.get("classification")),
            artifact_ref=_normalize_text(f2_report.get("artifact_paths", {}).get("plot_path")),
            track="candidate",
        )

    seen = {row["strategy_id"] for row in rows}
    for strategy_id, candidate in candidates.items():
        if strategy_id in seen:
            continue
        rows.append(
            {
                "strategy_id": strategy_id,
                "name": _display_strategy_name(strategy_id, candidate.get("name")),
                "role": _strategy_role(state, strategy_id, _normalize_text(candidate.get("track"))),
                "track": _normalize_text(candidate.get("track")),
                "decision": _normalize_text(candidate.get("decision")),
                "classification": "",
                "annualized_return": None,
                "max_drawdown": None,
                "sharpe_ratio": None,
                "calmar_ratio": None,
                "turnover": None,
                "win_rate": None,
                "artifact_ref": _first([str(item).strip() for item in candidate.get("artifact_refs", []) if str(item).strip()]),
            }
        )
    role_order = {"mainline": 0, "challenger": 1, "control": 2, "blocked": 3, "rejected": 4, "other": 5}
    rows.sort(key=lambda item: (role_order.get(str(item.get("role", "other")), 9), _normalize_text(item.get("strategy_id"))))
    return rows


def _preview_plot_rows(state: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for label, key in (
        ("F2 对照图", "f2_verify_report_path"),
        ("F1 对照图", "f1_verifier_report_path"),
        ("R1 对照图", "r1_verify_report_path"),
    ):
        report = _read_report(_normalize_text(state.get(key)))
        plot_path = _normalize_text(report.get("artifact_paths", {}).get("plot_path"))
        if not plot_path or plot_path in seen:
            continue
        seen.add(plot_path)
        rows.append({"label": label, "path": plot_path})
    return rows


def _strategies_rows(state: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for candidate in state.get("strategy_candidates", []) or []:
        if not isinstance(candidate, dict):
            continue
        artifact_refs = [str(item).strip() for item in candidate.get("artifact_refs", []) if str(item).strip()]
        rows.append(
            {
                "strategy_id": _normalize_text(candidate.get("strategy_id")),
                "name": _display_strategy_name(_normalize_text(candidate.get("strategy_id")), candidate.get("name")),
                "track": _normalize_text(candidate.get("track")),
                "current_stage": _normalize_text(candidate.get("current_stage")),
                "decision": _normalize_text(candidate.get("decision")),
                "latest_result": _normalize_text(candidate.get("latest_result")),
                "next_validation": _normalize_text(candidate.get("next_validation")),
                "artifact_ref": _first(artifact_refs),
            }
        )
    return rows


def _experiment_summary_rows(paths: ProjectPaths) -> list[dict[str, Any]]:
    counter: Counter[str] = Counter()
    for item in _read_jsonl(paths.experiment_ledger_path)[-100:]:
        counter[_normalize_text(item.get("result")) or "unknown"] += 1
    rows: list[dict[str, Any]] = []
    for classification, count in sorted(counter.items(), key=lambda pair: (-pair[1], pair[0])):
        rows.append({"classification": classification, "count": count})
    return rows


def _experiments_rows(paths: ProjectPaths) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, item in enumerate(list(reversed(_read_jsonl(paths.experiment_ledger_path)))[:100]):
        artifact_refs = [str(entry).strip() for entry in item.get("artifact_refs", []) if str(entry).strip()]
        rows.append(
            {
                "experiment_id": _normalize_text(item.get("experiment_id")),
                "timestamp": _normalize_text(item.get("timestamp")),
                "mode": _normalize_text(item.get("result")),
                "strategy_candidate_id": _normalize_text(item.get("strategy_candidate_id") or item.get("hypothesis")),
                "classification": _normalize_text(item.get("result")),
                "summary": _normalize_text(item.get("hypothesis")) or _normalize_text(item.get("result")),
                "report_path": _first(artifact_refs),
                "is_recent_key": "yes" if index < 15 else "no",
            }
        )
    rows.reverse()
    return rows


def _runs_rows(paths: ProjectPaths) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in list(reversed(_read_jsonl(paths.experiment_ledger_path)))[:100]:
        artifact_refs = [str(entry).strip() for entry in item.get("artifact_refs", []) if str(entry).strip()]
        blockers = [str(entry).strip() for entry in item.get("blockers", []) if str(entry).strip()]
        summary = _normalize_text(item.get("hypothesis")) or "; ".join(blockers) or _normalize_text(item.get("result"))
        rows.append(
            {
                "run_id": _normalize_text(item.get("experiment_id")),
                "kind": _normalize_text(item.get("result")),
                "status": _normalize_text(item.get("result")),
                "started_at": _normalize_text(item.get("timestamp")),
                "finished_at": _normalize_text(item.get("timestamp")),
                "summary": summary,
                "artifact_path": _first(artifact_refs),
            }
        )
    rows.reverse()
    return rows


def _strategy_role_summary_rows(strategy_metrics_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counts: Counter[str] = Counter()
    for row in strategy_metrics_rows:
        counts[_normalize_text(row.get("role")) or "other"] += 1
    order = ["mainline", "challenger", "control", "blocked", "rejected", "other"]
    return [
        {"role": role, "label": _role_label(role), "count": counts.get(role, 0)}
        for role in order
        if counts.get(role, 0) > 0
    ]


def _artifact_type(path: str) -> str:
    lowered = path.lower()
    if lowered.endswith(".png"):
        return "plot_png"
    if lowered.endswith(".md"):
        return "report_markdown"
    if lowered.endswith(".json"):
        return "report_json"
    if lowered.endswith(".csv"):
        return "table_csv"
    if lowered.endswith(".xlsx"):
        return "excel_console"
    if lowered.endswith((".cmd", ".bat", ".ps1", ".vbs", ".wsf", ".xlsm")):
        return "legacy_executable"
    return "artifact"


def _artifacts_rows(
    paths: ProjectPaths,
    state: dict[str, Any],
    console_paths: ExcelConsolePaths,
    manifest: dict[str, Any],
) -> list[dict[str, Any]]:
    seen: set[str] = set()
    refs: list[tuple[str, str]] = []
    for key, value in sorted(state.items()):
        if key.endswith("_report_path") and value:
            refs.append((key, str(value)))
    for candidate in state.get("strategy_candidates", []) or []:
        if not isinstance(candidate, dict):
            continue
        for ref in candidate.get("artifact_refs", []) or []:
            if ref:
                refs.append((str(candidate.get("strategy_id", "strategy_artifact")), str(ref)))
    for ledger_row in _read_jsonl(paths.experiment_ledger_path)[-50:]:
        for ref in ledger_row.get("artifact_refs", []) or []:
            if ref:
                refs.append((str(ledger_row.get("experiment_id", "experiment_artifact")), str(ref)))
    refs.extend(
        [
            ("project_state", str(paths.project_state_path)),
            ("research_memory", str(paths.research_memory_path)),
            ("strategy_board", str(paths.strategy_board_path)),
            ("verify_last", str(paths.verify_last_path)),
            ("handoff", str(paths.handoff_path)),
            ("migration_prompt", str(paths.migration_prompt_path)),
            ("session_state", str(paths.session_state_path)),
            ("excel_manifest", str(console_paths.manifest_path)),
            ("excel_console", str(Path(manifest.get("effective_workbook_path", console_paths.workbook_path)))),
        ]
    )
    rows: list[dict[str, Any]] = []
    for name, ref in refs:
        ref_text = _normalize_text(ref)
        if not ref_text or ref_text in seen:
            continue
        seen.add(ref_text)
        rows.append(
            {
                "artifact_type": _artifact_type(ref_text),
                "name": _normalize_text(name),
                "path": ref_text,
                "notes": "",
            }
        )
    return rows


def _columns_widths(rows: list[dict[str, Any]], columns: list[str]) -> dict[int, float]:
    widths: dict[int, float] = {}
    for index, column in enumerate(columns, start=1):
        max_len = len(column)
        for row in rows:
            max_len = max(max_len, len(str(row.get(column, ""))))
        widths[index] = max(12.0, min(float(max_len + 2), 60.0))
    return widths


def _style_header_row(ws: Worksheet, row_idx: int, columns: list[str], *, labels: dict[str, str] | None = None, start_col: int = 1) -> None:
    for index, column in enumerate(columns, start=1):
        cell = ws.cell(row=row_idx, column=start_col + index - 1, value=(labels or {}).get(column, column))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _BOX_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_path_hyperlink(cell, path_text: str) -> None:
    if not path_text:
        return
    cell.hyperlink = path_text
    cell.style = "Hyperlink"


def _write_metric_card(
    ws: Worksheet,
    *,
    start_cell: str,
    end_cell: str,
    title: str,
    value: str,
    fill: PatternFill = _CARD_FILL,
) -> None:
    ws.merge_cells(f"{start_cell}:{end_cell}")
    cell = ws[start_cell]
    cell.value = f"{title}\n{value}"
    cell.fill = fill
    cell.border = _BOX_BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    cell.font = Font(size=12, bold=True)


def _append_strategy_metric_table(
    ws: Worksheet,
    *,
    start_row: int,
    rows: list[dict[str, Any]],
    start_col: int = 1,
    columns: list[str] | None = None,
    labels: dict[str, str] | None = None,
) -> tuple[int, int]:
    columns = columns or [
        "strategy_id",
        "name",
        "role",
        "decision",
        "annualized_return",
        "max_drawdown",
        "sharpe_ratio",
        "calmar_ratio",
        "turnover",
    ]
    labels = labels or {
        "strategy_id": "策略 ID",
        "name": "展示名",
        "role": "角色",
        "decision": "当前结论",
        "annualized_return": "年化",
        "max_drawdown": "最大回撤",
        "sharpe_ratio": "Sharpe",
        "calmar_ratio": "Calmar",
        "turnover": "换手",
    }
    header_row = start_row
    _style_header_row(ws, header_row, columns, labels=labels, start_col=start_col)
    current_row = header_row + 1
    for row in rows:
        role = _normalize_text(row.get("role"))
        for index, column in enumerate(columns, start=1):
            raw_value = row.get(column)
            cell = ws.cell(row=current_row, column=start_col + index - 1)
            if column in {"annualized_return", "max_drawdown", "turnover"}:
                metric_value = _coerce_float(raw_value)
                cell.value = metric_value
                cell.number_format = "0.00%"
            elif column in {"sharpe_ratio", "calmar_ratio"}:
                metric_value = _coerce_float(raw_value)
                cell.value = metric_value
                cell.number_format = "0.000"
            elif column == "role":
                cell.value = _role_label(_normalize_text(raw_value))
            else:
                cell.value = raw_value
            cell.border = _BOX_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column == "decision":
                cell.fill = _card_fill_for_status(_normalize_text(raw_value))
            elif column in {"name", "role"}:
                cell.fill = _role_fill(role)
            if column == "name" and row.get("artifact_ref"):
                _apply_path_hyperlink(cell, str(row.get("artifact_ref")))
        current_row += 1
    default_widths = {
        "strategy_id": 26,
        "name": 18,
        "role": 10,
        "track": 12,
        "decision": 22,
        "current_stage": 16,
        "annualized_return": 12,
        "max_drawdown": 12,
        "sharpe_ratio": 10,
        "calmar_ratio": 10,
        "turnover": 10,
        "latest_result": 28,
        "next_validation": 28,
    }
    for offset, column in enumerate(columns, start=0):
        ws.column_dimensions[get_column_letter(start_col + offset)].width = default_widths.get(column, 16)
    return header_row, current_row - 1


def _add_strategy_comparison_chart(
    ws: Worksheet,
    *,
    data_start_row: int,
    data_end_row: int,
    anchor: str,
) -> None:
    if data_end_row <= data_start_row:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.height = 7
    chart.width = 12
    chart.title = "策略指标对比"
    chart.y_axis.title = "策略"
    chart.x_axis.title = "指标值"
    data = Reference(ws, min_col=5, min_row=data_start_row, max_col=8, max_row=data_end_row)
    categories = Reference(ws, min_col=2, min_row=data_start_row + 1, max_row=data_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.overlap = 0
    ws.add_chart(chart, anchor)


def _add_experiment_summary_chart(
    ws: Worksheet,
    *,
    start_row: int,
    rows: list[dict[str, Any]],
    anchor: str,
) -> None:
    if not rows:
        return
    ws.cell(row=start_row, column=1, value="result").font = _HEADER_FONT
    ws.cell(row=start_row, column=1).fill = _HEADER_FILL
    ws.cell(row=start_row, column=2, value="count").font = _HEADER_FONT
    ws.cell(row=start_row, column=2).fill = _HEADER_FILL
    current_row = start_row + 1
    for row in rows:
        ws.cell(row=current_row, column=1, value=row["classification"]).border = _BOX_BORDER
        ws.cell(row=current_row, column=2, value=row["count"]).border = _BOX_BORDER
        current_row += 1
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.height = 6
    chart.width = 9
    chart.title = "近期实验结果"
    data = Reference(ws, min_col=2, min_row=start_row, max_row=current_row - 1)
    categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=current_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, anchor)


def _add_role_summary_chart(
    ws: Worksheet,
    *,
    start_row: int,
    rows: list[dict[str, Any]],
    anchor: str,
) -> None:
    if not rows:
        return
    ws.cell(row=start_row, column=1, value="角色").font = _HEADER_FONT
    ws.cell(row=start_row, column=1).fill = _HEADER_FILL
    ws.cell(row=start_row, column=2, value="数量").font = _HEADER_FONT
    ws.cell(row=start_row, column=2).fill = _HEADER_FILL
    current_row = start_row + 1
    for row in rows:
        ws.cell(row=current_row, column=1, value=row["label"]).border = _BOX_BORDER
        ws.cell(row=current_row, column=2, value=row["count"]).border = _BOX_BORDER
        current_row += 1
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.height = 6
    chart.width = 8
    chart.title = "研究推进状态"
    data = Reference(ws, min_col=2, min_row=start_row, max_row=current_row - 1)
    categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=current_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, anchor)


def _add_preview_image(ws: Worksheet, *, image_path: str, anchor: str, target_width: int = 520) -> None:
    path = Path(image_path)
    if not path.exists():
        return
    image = XLImage(str(path))
    if image.width and image.height:
        ratio = target_width / float(image.width)
        image.width = target_width
        image.height = int(image.height * ratio)
    ws.add_image(image, anchor)


def _write_table_sheet(
    ws: Worksheet,
    *,
    title: str,
    rows: list[dict[str, Any]],
    columns: list[str],
    title_row: int = 1,
    start_row: int = 3,
    labels: dict[str, str] | None = None,
) -> None:
    title_cell = ws.cell(row=title_row, column=1, value=title)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left")
    _style_header_row(ws, start_row, columns, labels=labels)
    current_row = start_row + 1
    path_like_columns = {"artifact_ref", "report_path", "artifact_path", "path"}
    for row in rows:
        for index, column in enumerate(columns, start=1):
            value = row.get(column, "")
            cell = ws.cell(row=current_row, column=index, value=value)
            cell.border = _BOX_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column == "role":
                cell.fill = _role_fill(_normalize_text(row.get("role")))
                cell.value = _role_label(_normalize_text(value))
            if column == "decision":
                cell.fill = _card_fill_for_status(_normalize_text(value))
            if column in path_like_columns and value:
                _apply_path_hyperlink(cell, str(value))
        current_row += 1
    end_column = get_column_letter(len(columns))
    end_row = max(start_row + 1, current_row - 1)
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{end_column}{end_row}"
    for index, width in _columns_widths(rows, columns).items():
        ws.column_dimensions[get_column_letter(index)].width = width


def _write_overview_sheet(
    ws: Worksheet,
    research_summary_rows: list[dict[str, Any]],
    *,
    strategy_metrics_rows: list[dict[str, Any]],
    experiment_summary_rows: list[dict[str, Any]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "研究总览"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left")
    mainline = _strategy_lookup_by_id(strategy_metrics_rows, "f1_elasticnet_v1")
    challenger = _strategy_lookup_by_id(strategy_metrics_rows, "f2_structured_latent_factor_v1")
    _write_metric_card(
        ws,
        start_cell="A3",
        end_cell="B5",
        title="当前主线",
        value=_display_strategy_name("f1_elasticnet_v1", mainline.get("name")),
        fill=_SUCCESS_FILL,
    )
    _write_metric_card(
        ws,
        start_cell="C3",
        end_cell="D5",
        title="F1 年化",
        value=_fmt_pct(mainline.get("annualized_return")) or "n/a",
        fill=_CARD_FILL,
    )
    _write_metric_card(
        ws,
        start_cell="E3",
        end_cell="F5",
        title="F2 年化",
        value=_fmt_pct(challenger.get("annualized_return")) or "n/a",
        fill=_CARD_FILL,
    )
    _write_metric_card(
        ws,
        start_cell="G3",
        end_cell="H5",
        title="当前 blocker",
        value=_first(item["value"] for item in research_summary_rows if item["key"] == "当前 blocker") or "n/a",
        fill=_CAUTION_FILL,
    )
    ws["A7"] = "研究状态摘要"
    ws["A7"].font = _SECTION_TITLE_FONT
    ws["A7"].fill = _HEADER_FILL
    ws["A7"].border = _BOX_BORDER
    ws.merge_cells("A7:B7")
    row = 8
    for item in sorted(research_summary_rows, key=lambda entry: int(entry["display_order"])):
        key_cell = ws.cell(row=row, column=1, value=str(item["key"]))
        value_cell = ws.cell(row=row, column=2, value=str(item["value"]))
        key_cell.font = _LABEL_FONT
        key_cell.border = _BOX_BORDER
        value_cell.border = _BOX_BORDER
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    _add_role_summary_chart(
        ws,
        start_row=8,
        rows=_strategy_role_summary_rows(strategy_metrics_rows),
        anchor="E7",
    )
    _add_experiment_summary_chart(ws, start_row=max(row + 2, 18), rows=experiment_summary_rows, anchor="E18")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["E"].width = 16
    ws.freeze_panes = "A8"


def _safe_command_palette(project: str, *, console_paths: ExcelConsolePaths) -> list[dict[str, str]]:
    artifacts_dir = console_paths.root.parent
    tracked_memory_dir = console_paths.root.parents[3] / "memory" / "projects" / project
    return [
        {"label": "刷新数据包", "command": f"python -m quant_mvp excel_export --project {project}"},
        {"label": "运行 research_audit", "command": f"python -m quant_mvp research_audit --project {project}"},
        {"label": "运行 agent_cycle --dry-run", "command": f"python -m quant_mvp agent_cycle --project {project} --dry-run"},
        {"label": "运行 f1_verify", "command": f"python -m quant_mvp f1_verify --project {project}"},
        {"label": "运行 f2_verify", "command": f"python -m quant_mvp f2_verify --project {project}"},
        {"label": "打开产物目录", "command": f'explorer "{artifacts_dir}"'},
        {"label": "打开 tracked memory", "command": f'explorer "{tracked_memory_dir}"'},
        {"label": "打开当前工作簿", "command": f'explorer "{console_paths.workbook_path}"'},
    ]


def _write_command_row(ws: Worksheet, *, row: int, label: str, command: str) -> None:
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=14)
    label_cell = ws.cell(row=row, column=7, value=label)
    command_cell = ws.cell(row=row, column=9, value=command)
    label_cell.fill = _BUTTON_FILL
    label_cell.font = Font(color="0B5394", bold=True)
    label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    command_cell.fill = _CARD_FILL
    command_cell.font = Font(name="Consolas", size=10, color="1F1F1F")
    command_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col in range(7, 15):
        ws.cell(row=row, column=col).border = _BOX_BORDER


def _write_control_sheet(
    ws: Worksheet,
    *,
    project: str,
    state: dict[str, Any],
    manifest: dict[str, Any],
    safe_commands: list[dict[str, str]],
    control_cards_rows: list[dict[str, Any]],
    strategy_metrics_rows: list[dict[str, Any]],
    experiment_summary_rows: list[dict[str, Any]],
    preview_plot_rows: list[dict[str, str]],
) -> None:
    del experiment_summary_rows
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:N1")
    ws["A1"] = "研究主控台 / Research Console"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.merge_cells("A2:N2")
    ws["A2"] = f"项目：{project}"
    ws["A2"].font = _SMALL_MUTED_FONT
    ws["A2"].alignment = Alignment(horizontal="left")

    primary = _strategy_lookup_by_id(strategy_metrics_rows, "f1_elasticnet_v1")
    challenger = _strategy_lookup_by_id(strategy_metrics_rows, "f2_structured_latent_factor_v1")
    control = _strategy_lookup_by_id(strategy_metrics_rows, "baseline_limit_up")
    cards = {row["slot"]: row for row in control_cards_rows}
    focus_rows = [
        row
        for row in strategy_metrics_rows
        if row["strategy_id"] in {"baseline_limit_up", "f1_elasticnet_v1", "f2_structured_latent_factor_v1"}
    ]

    _write_metric_card(
        ws,
        start_cell="A4",
        end_cell="C6",
        title=str(cards.get("card_mainline", {}).get("title") or "当前主线"),
        value=str(cards.get("card_mainline", {}).get("value") or _display_strategy_name("f1_elasticnet_v1", primary.get("name")) or "n/a"),
        fill=_SUCCESS_FILL,
    )
    _write_metric_card(
        ws,
        start_cell="D4",
        end_cell="G6",
        title=str(cards.get("card_blocker", {}).get("title") or "当前 blocker"),
        value=str(cards.get("card_blocker", {}).get("value") or _normalize_text(state.get("current_blocker")) or "n/a"),
        fill=_CAUTION_FILL,
    )
    _write_metric_card(
        ws,
        start_cell="H4",
        end_cell="K6",
        title=str(cards.get("card_next_step", {}).get("title") or "下一步"),
        value=str(cards.get("card_next_step", {}).get("value") or _normalize_text(state.get("next_priority_action")) or "n/a"),
        fill=_SUBTLE_FILL,
    )
    _write_metric_card(
        ws,
        start_cell="L4",
        end_cell="N6",
        title=str(cards.get("card_conclusion", {}).get("title") or "当前结论"),
        value=str(cards.get("card_conclusion", {}).get("value") or _latest_conclusion_text(state) or "n/a"),
        fill=_CARD_FILL,
    )

    _write_metric_card(
        ws,
        start_cell="A8",
        end_cell="C9",
        title=str(cards.get("metric_f1", {}).get("title") or "F1 核心指标"),
        value=str(cards.get("metric_f1", {}).get("value") or f"{_fmt_pct(primary.get('annualized_return')) or 'n/a'} / {_fmt_pct(primary.get('max_drawdown')) or 'n/a'}"),
        fill=_SUCCESS_FILL,
    )
    _write_metric_card(
        ws,
        start_cell="D8",
        end_cell="F9",
        title=str(cards.get("metric_f2", {}).get("title") or "F2 核心指标"),
        value=str(cards.get("metric_f2", {}).get("value") or f"{_fmt_pct(challenger.get('annualized_return')) or 'n/a'} / {_fmt_pct(challenger.get('max_drawdown')) or 'n/a'}"),
        fill=_card_fill_for_status(_normalize_text(challenger.get("decision"))),
    )
    _write_metric_card(
        ws,
        start_cell="G8",
        end_cell="I9",
        title=str(cards.get("metric_control", {}).get("title") or "对照指标"),
        value=str(cards.get("metric_control", {}).get("value") or f"{_fmt_pct(control.get('annualized_return')) or 'n/a'} / {_fmt_pct(control.get('max_drawdown')) or 'n/a'}"),
        fill=_NEUTRAL_FILL,
    )
    _write_metric_card(
        ws,
        start_cell="J8",
        end_cell="N9",
        title=str(cards.get("metric_export", {}).get("title") or "最近导出时间"),
        value=str(cards.get("metric_export", {}).get("value") or _normalize_text(manifest.get("generated_at")) or "n/a"),
        fill=_CARD_FILL,
    )

    ws["A11"] = "最近验证能力"
    ws["A11"].font = _SECTION_TITLE_FONT
    ws["A11"].fill = _HEADER_FILL
    ws["A11"].border = _BOX_BORDER
    ws.merge_cells("A11:F11")
    ws["A12"] = (
        f"最新结论：{_latest_conclusion_text(state)}\n"
        f"最新验证能力：{_normalize_text(state.get('last_verified_capability'))}\n"
        f"验证口径：{_research_scope_text(state)}"
    )
    ws["A12"].fill = _SUBTLE_FILL
    ws["A12"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A12:F15")

    ws["G11"] = "安全命令（复制到终端）"
    ws["G11"].font = _SECTION_TITLE_FONT
    ws["G11"].fill = _HEADER_FILL
    ws["G11"].border = _BOX_BORDER
    ws.merge_cells("G11:N11")
    for offset, item in enumerate(safe_commands, start=12):
        _write_command_row(ws, row=offset, label=item["label"], command=item["command"])

    ws["A21"] = "主线对照图"
    ws["A21"].font = _SECTION_TITLE_FONT
    ws["A21"].fill = _HEADER_FILL
    ws["A21"].border = _BOX_BORDER
    ws.merge_cells("A21:H21")
    if preview_plot_rows:
        _add_preview_image(ws, image_path=preview_plot_rows[0]["path"], anchor="A22", target_width=860)

    ws["J21"] = "主线对照表"
    ws["J21"].font = _SECTION_TITLE_FONT
    ws["J21"].fill = _HEADER_FILL
    ws["J21"].border = _BOX_BORDER
    ws.merge_cells("J21:N21")
    _append_strategy_metric_table(
        ws,
        start_row=22,
        start_col=10,
        rows=focus_rows,
        columns=["name", "role", "decision", "annualized_return", "max_drawdown", "sharpe_ratio"],
        labels={
            "name": "策略",
            "role": "角色",
            "decision": "结论",
            "annualized_return": "年化",
            "max_drawdown": "最大回撤",
            "sharpe_ratio": "Sharpe",
        },
    )

    for row_idx, height in {
        1: 28,
        2: 18,
        4: 40,
        8: 34,
        12: 28,
        13: 28,
        14: 28,
        15: 28,
    }.items():
        ws.row_dimensions[row_idx].height = height
    for column, width in {
        "A": 16,
        "B": 16,
        "C": 16,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 18,
        "H": 18,
        "I": 16,
        "J": 16,
        "K": 16,
        "L": 16,
        "M": 16,
        "N": 16,
    }.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A11"


def _write_strategies_sheet(
    ws: Worksheet,
    *,
    strategies_rows: list[dict[str, Any]],
    strategy_metrics_rows: list[dict[str, Any]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:I1")
    ws["A1"] = "策略面板"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left")
    ws["A3"] = "核心指标对比"
    ws["A3"].font = _SECTION_TITLE_FONT
    ws["A3"].fill = _HEADER_FILL
    ws["A3"].border = _BOX_BORDER
    ws.merge_cells("A3:I3")
    header_row, end_row = _append_strategy_metric_table(ws, start_row=4, rows=strategy_metrics_rows)
    _add_strategy_comparison_chart(ws, data_start_row=header_row, data_end_row=end_row, anchor="K3")
    _write_table_sheet(
        ws,
        title="策略账本",
        rows=strategies_rows,
        columns=["strategy_id", "name", "track", "current_stage", "decision", "latest_result", "next_validation", "artifact_ref"],
        title_row=end_row + 2,
        start_row=end_row + 4,
        labels={
            "strategy_id": "策略 ID",
            "name": "展示名",
            "track": "轨道",
            "current_stage": "当前阶段",
            "decision": "当前结论",
            "latest_result": "最新结果",
            "next_validation": "下一验证",
            "artifact_ref": "报告路径",
        },
    )
    ws.freeze_panes = "A4"


def _write_experiments_sheet(
    ws: Worksheet,
    *,
    experiments_rows: list[dict[str, Any]],
    experiment_summary_rows: list[dict[str, Any]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "实验账本"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left")
    _add_experiment_summary_chart(ws, start_row=3, rows=experiment_summary_rows, anchor="D3")
    recent_key_rows = [row for row in experiments_rows if row.get("is_recent_key") == "yes"][:15]
    _write_table_sheet(
        ws,
        title="近期关键实验",
        rows=recent_key_rows,
        columns=["experiment_id", "timestamp", "strategy_candidate_id", "classification", "summary", "report_path"],
        title_row=15,
        start_row=17,
        labels={
            "experiment_id": "实验 ID",
            "timestamp": "时间",
            "strategy_candidate_id": "对象",
            "classification": "分类",
            "summary": "摘要",
            "report_path": "报告路径",
        },
    )
    _write_table_sheet(
        ws,
        title="完整实验流水",
        rows=experiments_rows,
        columns=["experiment_id", "timestamp", "mode", "strategy_candidate_id", "classification", "summary", "report_path"],
        title_row=35,
        start_row=37,
        labels={
            "experiment_id": "实验 ID",
            "timestamp": "时间",
            "mode": "模式",
            "strategy_candidate_id": "对象",
            "classification": "分类",
            "summary": "摘要",
            "report_path": "报告路径",
        },
    )
    ws.freeze_panes = "A4"


def _write_artifacts_sheet(
    ws: Worksheet,
    *,
    artifacts_rows: list[dict[str, Any]],
    preview_plot_rows: list[dict[str, str]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "报告与图表"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left")
    for index, preview in enumerate(preview_plot_rows[:3]):
        anchor_row = 3 + index * 18
        ws.cell(row=anchor_row, column=1, value=preview["label"]).font = _SECTION_TITLE_FONT
        ws.cell(row=anchor_row, column=1).fill = _HEADER_FILL
        ws.cell(row=anchor_row, column=1).border = _BOX_BORDER
        ws.merge_cells(start_row=anchor_row, start_column=1, end_row=anchor_row, end_column=8)
        _add_preview_image(ws, image_path=preview["path"], anchor=f"A{anchor_row + 1}", target_width=820)
    _write_table_sheet(
        ws,
        title="产物索引",
        rows=artifacts_rows,
        columns=["artifact_type", "name", "path", "notes"],
        title_row=58,
        start_row=60,
        labels={
            "artifact_type": "产物类型",
            "name": "名称",
            "path": "路径",
            "notes": "备注",
        },
    )


def _write_notes(console_paths: ExcelConsolePaths, *, manifest: dict[str, Any]) -> Path:
    macro_status = str(manifest.get("macro_injection_status", "unknown"))
    workbook_mode = str(manifest.get("workbook_mode", "unknown"))
    removed_legacy = manifest.get("removed_legacy_files", []) or []
    text = "\n".join(
        [
            "# Excel 控制台说明",
            "",
            f"- 导出时间: {manifest.get('generated_at', '')}",
            f"- 工作簿模式: {workbook_mode}",
            f"- 宏状态: {macro_status}",
            "- 首页是驾驶舱，不是全量台账。",
            "- 中文优先，Python 仍是唯一真相源。",
            "- 当前工作簿只生成安全的 `.xlsx` 仪表盘，不再生成 `.cmd`、`.bat`、`.ps1`、VBA 模块或任何可执行启动器脚本。",
            "- 首页动作区只展示可复制的终端命令；真正执行仍在终端完成。",
            f"- 本次已清理的旧脚本/旧工作簿数量: {len(removed_legacy)}",
            "",
            "## 冻结状态",
            "- `apps/web` 继续冻结，不删除。",
            "- `dashboard/app.py` 继续冻结，不删除。",
            "- 只有当 Excel 已覆盖项目概览、实验账本、最近验证结果和安全命令后，才允许删除网页入口。",
        ]
    )
    return _write_text(console_paths.notes_path, text)


def _build_workbook(
    *,
    workbook_path: Path,
    project: str,
    state: dict[str, Any],
    manifest: dict[str, Any],
    safe_commands: list[dict[str, str]],
    overview_rows: list[dict[str, Any]],
    research_summary_rows: list[dict[str, Any]],
    control_cards_rows: list[dict[str, Any]],
    strategies_rows: list[dict[str, Any]],
    strategy_metrics_rows: list[dict[str, Any]],
    experiment_summary_rows: list[dict[str, Any]],
    experiments_rows: list[dict[str, Any]],
    runs_rows: list[dict[str, Any]],
    artifacts_rows: list[dict[str, Any]],
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    control_ws = workbook.create_sheet(WORKSHEET_NAMES[0])
    overview_ws = workbook.create_sheet(WORKSHEET_NAMES[1])
    strategies_ws = workbook.create_sheet(WORKSHEET_NAMES[2])
    experiments_ws = workbook.create_sheet(WORKSHEET_NAMES[3])
    runs_ws = workbook.create_sheet(WORKSHEET_NAMES[4])
    artifacts_ws = workbook.create_sheet(WORKSHEET_NAMES[5])
    preview_plot_rows = _preview_plot_rows(state)

    _write_control_sheet(
        control_ws,
        project=project,
        state=state,
        manifest=manifest,
        safe_commands=safe_commands,
        control_cards_rows=control_cards_rows,
        strategy_metrics_rows=strategy_metrics_rows,
        experiment_summary_rows=experiment_summary_rows,
        preview_plot_rows=preview_plot_rows,
    )
    _write_overview_sheet(
        overview_ws,
        research_summary_rows,
        strategy_metrics_rows=strategy_metrics_rows,
        experiment_summary_rows=experiment_summary_rows,
    )
    _write_strategies_sheet(
        strategies_ws,
        strategies_rows=strategies_rows,
        strategy_metrics_rows=strategy_metrics_rows,
    )
    _write_experiments_sheet(
        experiments_ws,
        experiments_rows=experiments_rows,
        experiment_summary_rows=experiment_summary_rows,
    )
    _write_table_sheet(
        runs_ws,
        title="运行记录",
        rows=runs_rows,
        columns=["run_id", "kind", "status", "started_at", "finished_at", "summary", "artifact_path"],
        labels={
            "run_id": "运行 ID",
            "kind": "类型",
            "status": "状态",
            "started_at": "开始时间",
            "finished_at": "结束时间",
            "summary": "摘要",
            "artifact_path": "产物路径",
        },
    )
    _write_artifacts_sheet(
        artifacts_ws,
        artifacts_rows=artifacts_rows,
        preview_plot_rows=preview_plot_rows,
    )
    staging_dir = workbook_path.parent / "_staging"
    staging_dir.mkdir(parents=True, exist_ok=True)
    temp_xlsx = staging_dir / f"{workbook_path.stem}__staging.xlsx"
    workbook.save(temp_xlsx)
    workbook_path.unlink(missing_ok=True)
    shutil.move(str(temp_xlsx), str(workbook_path))
    shutil.rmtree(staging_dir, ignore_errors=True)
    return workbook_path


def _write_manifest(console_paths: ExcelConsolePaths, manifest: dict[str, Any]) -> Path:
    console_paths.manifest_path.parent.mkdir(parents=True, exist_ok=True)
    console_paths.manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True).rstrip() + "\n",
        encoding="utf-8",
    )
    return console_paths.manifest_path


def run_excel_export(
    project: str,
    *,
    config_path: Path | None = None,
    repo_root: Path | None = None,
    probe_vba: bool = True,
) -> dict[str, Any]:
    root = find_repo_root(repo_root)
    requested_config_path = ""
    if config_path is not None:
        resolved_config = config_path.resolve()
        if not resolved_config.exists():
            raise FileNotFoundError(f"Config path not found: {resolved_config}")
        requested_config_path = str(resolved_config)
    bootstrap_memory_files(project, repo_root=root)
    paths, state = load_machine_state(project, repo_root=root)
    console_paths = _excel_console_paths(paths)
    _ensure_excel_dirs(console_paths)
    removed_legacy_files = _cleanup_legacy_excel_artifacts(console_paths)

    vba_status = {
        "excel_com_available": False,
        "vbproject_access": False,
        "status": "disabled_no_scripts",
        "detail": "Excel console no longer emits executable launcher scripts or VBA automation.",
    }
    workbook_mode = "safe_dashboard_no_scripts"
    generated_at = _utc_now()
    safe_commands = _safe_command_palette(paths.project, console_paths=console_paths)

    overview_rows = _overview_rows(state, workbook_mode=workbook_mode, workbook_path=console_paths.workbook_path)
    strategy_metrics_rows = _strategy_metrics_rows(state)
    research_summary_rows = _research_summary_rows(state)
    control_cards_rows = _control_cards_rows(
        state,
        strategy_metrics_rows,
        exported_at=generated_at,
    )
    strategies_rows = _strategies_rows(state)
    experiment_summary_rows = _experiment_summary_rows(paths)
    experiments_rows = _experiments_rows(paths)
    runs_rows = _runs_rows(paths)

    feed_files = {
        "overview": _write_csv(
            console_paths.feed_dir / "overview.csv",
            overview_rows,
            ["section", "key", "value", "display_order"],
        ),
        "research_summary": _write_csv(
            console_paths.feed_dir / "research_summary.csv",
            research_summary_rows,
            ["section", "key", "value", "display_order"],
        ),
        "control_cards": _write_csv(
            console_paths.feed_dir / "control_cards.csv",
            control_cards_rows,
            ["slot", "title", "value", "status", "display_order"],
        ),
        "strategies": _write_csv(
            console_paths.feed_dir / "strategies.csv",
            strategies_rows,
            ["strategy_id", "name", "track", "current_stage", "decision", "latest_result", "next_validation", "artifact_ref"],
        ),
        "strategy_metrics": _write_csv(
            console_paths.feed_dir / "strategy_metrics.csv",
            strategy_metrics_rows,
            [
                "strategy_id",
                "name",
                "role",
                "track",
                "decision",
                "classification",
                "annualized_return",
                "max_drawdown",
                "sharpe_ratio",
                "calmar_ratio",
                "turnover",
                "win_rate",
                "artifact_ref",
            ],
        ),
        "experiments": _write_csv(
            console_paths.feed_dir / "experiments.csv",
            experiments_rows,
            ["experiment_id", "timestamp", "mode", "strategy_candidate_id", "classification", "summary", "report_path", "is_recent_key"],
        ),
        "experiment_summary": _write_csv(
            console_paths.feed_dir / "experiment_summary.csv",
            experiment_summary_rows,
            ["classification", "count"],
        ),
        "runs": _write_csv(
            console_paths.feed_dir / "runs.csv",
            runs_rows,
            ["run_id", "kind", "status", "started_at", "finished_at", "summary", "artifact_path"],
        ),
    }
    manifest_stub = {
        "project": paths.project,
        "generated_at": generated_at,
        "feed_version": FEED_VERSION,
        "workbook_mode": workbook_mode,
        "macro_injection_status": str(vba_status.get("status", "unknown")),
        "macro_probe": vba_status,
        "primary_workbook_path": str(console_paths.workbook_path),
        "requested_config_path": requested_config_path,
        "config_override_mode": "tracked_truth_export_only",
        "removed_legacy_files": removed_legacy_files,
    }
    artifacts_rows = _artifacts_rows(paths, state, console_paths, manifest_stub)
    feed_files["artifacts"] = _write_csv(
        console_paths.feed_dir / "artifacts.csv",
        artifacts_rows,
        ["artifact_type", "name", "path", "notes"],
    )

    workbook_output_path = console_paths.workbook_path
    workbook_write_status = "primary_workbook_updated"
    try:
        _build_workbook(
            workbook_path=console_paths.workbook_path,
            project=paths.project,
            state=state,
            manifest={**manifest_stub, "effective_workbook_path": str(console_paths.workbook_path)},
            safe_commands=safe_commands,
            overview_rows=overview_rows,
            research_summary_rows=research_summary_rows,
            control_cards_rows=control_cards_rows,
            strategies_rows=strategies_rows,
            strategy_metrics_rows=strategy_metrics_rows,
            experiment_summary_rows=experiment_summary_rows,
            experiments_rows=experiments_rows,
            runs_rows=runs_rows,
            artifacts_rows=artifacts_rows,
        )
    except PermissionError:
        workbook_output_path = console_paths.root / f"ResearchConsole_{datetime.now(UTC).strftime('%Y%m%dT%H%M%SZ')}.xlsx"
        workbook_write_status = "primary_workbook_locked_wrote_timestamped_copy"
        _build_workbook(
            workbook_path=workbook_output_path,
            project=paths.project,
            state=state,
            manifest={**manifest_stub, "effective_workbook_path": str(workbook_output_path)},
            safe_commands=safe_commands,
            overview_rows=overview_rows,
            research_summary_rows=research_summary_rows,
            control_cards_rows=control_cards_rows,
            strategies_rows=strategies_rows,
            strategy_metrics_rows=strategy_metrics_rows,
            experiment_summary_rows=experiment_summary_rows,
            experiments_rows=experiments_rows,
            runs_rows=runs_rows,
            artifacts_rows=artifacts_rows,
        )

    manifest = {
        "project": paths.project,
        "generated_at": _utc_now(),
        "feed_version": FEED_VERSION,
        "workbook_mode": workbook_mode,
        "macro_injection_status": str(vba_status.get("status", "unknown")),
        "macro_probe": vba_status,
        "primary_workbook_path": str(console_paths.workbook_path),
        "requested_config_path": requested_config_path,
        "config_override_mode": "tracked_truth_export_only",
        "effective_workbook_path": str(workbook_output_path),
        "workbook_write_status": workbook_write_status,
        "feed_files": {key: str(value) for key, value in feed_files.items()},
        "command_palette": safe_commands,
        "notes_path": str(console_paths.notes_path),
        "web_freeze_status": "frozen_until_excel_mvp_acceptance",
        "safe_actions": [item["label"] for item in safe_commands],
        "removed_legacy_files": removed_legacy_files,
    }
    artifacts_rows = _artifacts_rows(paths, state, console_paths, manifest)
    feed_files["artifacts"] = _write_csv(
        console_paths.feed_dir / "artifacts.csv",
        artifacts_rows,
        ["artifact_type", "name", "path", "notes"],
    )
    _write_manifest(console_paths, manifest)
    _write_notes(console_paths, manifest=manifest)
    _build_workbook(
        workbook_path=workbook_output_path,
        project=paths.project,
        state=state,
        manifest=manifest,
        safe_commands=safe_commands,
        overview_rows=overview_rows,
        research_summary_rows=research_summary_rows,
        control_cards_rows=control_cards_rows,
        strategies_rows=strategies_rows,
        strategy_metrics_rows=strategy_metrics_rows,
        experiment_summary_rows=experiment_summary_rows,
        experiments_rows=experiments_rows,
        runs_rows=runs_rows,
        artifacts_rows=artifacts_rows,
    )

    return {
        "project": paths.project,
        "excel_root": str(console_paths.root),
        "workbook_path": str(workbook_output_path),
        "manifest_path": str(console_paths.manifest_path),
        "workbook_mode": workbook_mode,
        "macro_injection_status": manifest["macro_injection_status"],
        "feed_files": {key: str(value) for key, value in feed_files.items()},
        "command_palette": safe_commands,
    }
