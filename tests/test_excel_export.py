from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from quant_mvp.excel_export import FEED_VERSION, WORKSHEET_NAMES, run_excel_export
from quant_mvp.memory.writeback import bootstrap_memory_files, load_machine_state, record_experiment_result, save_machine_state


def _seed_console_state(project: str, *, repo_root: Path) -> None:
    bootstrap_memory_files(project, repo_root=repo_root)
    paths, state = load_machine_state(project, repo_root=repo_root)
    f1_report_path = paths.artifacts_dir / "f1" / "F1_BOUNDED_VERIFIER.json"
    f1_report_path.parent.mkdir(parents=True, exist_ok=True)
    f1_report_path.write_text(
        json.dumps(
            {
                "decision": "keep_f1_mainline",
                "classification": "verifier_pass",
                "f1_metrics": {"annualized_return": 0.5273, "max_drawdown": -0.3710, "sharpe_ratio": 1.68, "calmar_ratio": 1.42},
                "control_metrics": {"annualized_return": 0.0053, "max_drawdown": -0.6532, "sharpe_ratio": -0.073, "calmar_ratio": 0.008},
                "artifact_paths": {"plot_path": str(paths.artifacts_dir / "f1" / "f1_vs_control_common_shell.png")},
            },
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    f2_report_path = paths.artifacts_dir / "f2" / "F2_BOUNDED_VERIFIER.json"
    f2_report_path.parent.mkdir(parents=True, exist_ok=True)
    f2_report_path.write_text(
        json.dumps(
            {
                "decision": "keep_f2_challenger",
                "classification": "verifier_mixed",
                "f2_metrics": {"annualized_return": 0.5347, "max_drawdown": -0.3358, "sharpe_ratio": 1.586, "calmar_ratio": 1.59},
                "artifact_paths": {"plot_path": str(paths.artifacts_dir / "f2" / "f2_vs_f1_vs_control.png")},
            },
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )

    state.update(
        {
            "project": project,
            "current_phase": "F2.1 bounded challenger",
            "current_task": "Keep F1 as the verified mainline and review one more bounded F2 variant.",
            "current_blocker": "F2.1 shared-shell Top6 max_drawdown 33.58% remains above 30.00%.",
            "current_conclusion": "F2.1 已跑通，但当前仍只是挑战者，尚未取代 F1 主线。",
            "current_validation_scope": "同一核心股票池 + 同一 shared shell + Top6。",
            "next_priority_action": "再做一个 bounded F2 变体，不扩搜索范围。",
            "last_verified_capability": "F2.1 verifier completed under the shared shell and kept F2.1 as a challenger.",
            "canonical_universe_id": "cn_a_mainboard_all_v1",
            "effective_subagent_gate_mode": "OFF",
            "readiness": {"ready": True, "stage": "validation-ready"},
            "current_primary_strategy_ids": ["f1_elasticnet_v1"],
            "current_secondary_strategy_ids": ["f2_structured_latent_factor_v1", "baseline_limit_up"],
            "current_blocked_strategy_ids": ["risk_constrained_limit_up"],
            "f1_verifier_report_path": str(f1_report_path),
            "f2_verify_report_path": str(f2_report_path),
            "f2_verify_decision": "keep_f2_challenger",
            "strategy_candidates": [
                {
                    "strategy_id": "f1_elasticnet_v1",
                    "name": "F1 主线",
                    "track": "primary",
                    "current_stage": "validation",
                    "decision": "keep_f1_mainline",
                    "latest_result": "F1 shared-shell Top6: annualized=52.73%, drawdown=37.10%.",
                    "next_validation": "Keep F1 as mainline while testing one more bounded F2 variant.",
                    "artifact_refs": [str(f1_report_path)],
                },
                {
                    "strategy_id": "f2_structured_latent_factor_v1",
                    "name": "F2 挑战者",
                    "track": "candidate",
                    "current_stage": "verifier_mixed",
                    "decision": "keep_f2_challenger",
                    "latest_result": "F2.1 improved the tradeoff but did not displace F1.",
                    "next_validation": "Run one more bounded F2 variant under the same shell.",
                    "artifact_refs": [str(f2_report_path)],
                },
                {
                    "strategy_id": "baseline_limit_up",
                    "name": "基线对照",
                    "track": "control",
                    "current_stage": "control",
                    "decision": "continue",
                    "latest_result": "Kept only as a control harness.",
                    "next_validation": "Use the same shared shell for comparisons.",
                    "artifact_refs": [str(f1_report_path)],
                },
            ],
        }
    )
    save_machine_state(project, state, repo_root=repo_root)
    record_experiment_result(
        project,
        {
            "timestamp": "2026-03-31T10:38:56+00:00",
            "experiment_id": f"{project}__factor_elasticnet_core__f2_verify__20260331T103833Z",
            "hypothesis": "F2.1 should improve F1's tradeoff on the same shared shell.",
            "result": "verifier_mixed",
            "artifact_refs": [str(f2_report_path)],
            "blockers": ["drawdown still above threshold"],
        },
        repo_root=repo_root,
    )


def test_excel_export_generates_feed_and_workbook(synthetic_project) -> None:
    project = synthetic_project["project"]
    repo_root = synthetic_project["paths"].root
    _seed_console_state(project, repo_root=repo_root)

    result = run_excel_export(project, repo_root=repo_root, probe_vba=False)

    manifest_path = Path(result["manifest_path"])
    workbook_path = Path(result["workbook_path"])
    assert manifest_path.exists()
    assert workbook_path.exists()

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["project"] == project
    assert manifest["feed_version"] == FEED_VERSION
    assert manifest["workbook_mode"] == "safe_dashboard_no_scripts"
    assert manifest["macro_injection_status"] == "disabled_no_scripts"
    assert "removed_legacy_files" in manifest

    overview = pd.read_csv(Path(result["feed_files"]["overview"]))
    research_summary = pd.read_csv(Path(result["feed_files"]["research_summary"]))
    control_cards = pd.read_csv(Path(result["feed_files"]["control_cards"]))
    strategies = pd.read_csv(Path(result["feed_files"]["strategies"]))
    strategy_metrics = pd.read_csv(Path(result["feed_files"]["strategy_metrics"]))
    experiments = pd.read_csv(Path(result["feed_files"]["experiments"]))
    experiment_summary = pd.read_csv(Path(result["feed_files"]["experiment_summary"]))
    runs = pd.read_csv(Path(result["feed_files"]["runs"]))
    artifacts = pd.read_csv(Path(result["feed_files"]["artifacts"]))

    assert list(overview.columns) == ["section", "key", "value", "display_order"]
    assert list(research_summary.columns) == ["section", "key", "value", "display_order"]
    assert list(control_cards.columns) == ["slot", "title", "value", "status", "display_order"]
    assert list(strategies.columns) == [
        "strategy_id",
        "name",
        "track",
        "current_stage",
        "decision",
        "latest_result",
        "next_validation",
        "artifact_ref",
    ]
    assert list(strategy_metrics.columns) == [
        "strategy_id",
        "name",
        "role",
        "track",
        "decision",
        "classification",
        "annualized_return",
        "max_drawdown",
        "sharpe_ratio",
        "calmar_ratio",
        "turnover",
        "win_rate",
        "artifact_ref",
    ]
    assert list(experiments.columns) == [
        "experiment_id",
        "timestamp",
        "mode",
        "strategy_candidate_id",
        "classification",
        "summary",
        "report_path",
        "is_recent_key",
    ]
    assert list(experiment_summary.columns) == ["classification", "count"]
    assert list(runs.columns) == ["run_id", "kind", "status", "started_at", "finished_at", "summary", "artifact_path"]
    assert list(artifacts.columns) == ["artifact_type", "name", "path", "notes"]
    assert set(strategies["strategy_id"]) == {"f1_elasticnet_v1", "f2_structured_latent_factor_v1", "baseline_limit_up"}
    assert "当前 blocker" in set(research_summary["key"])
    assert "card_mainline" in set(control_cards["slot"])

    workbook = load_workbook(workbook_path)
    assert tuple(workbook.sheetnames) == WORKSHEET_NAMES
    control = workbook["Control"]
    assert control["A1"].value == "研究主控台 / Research Console"
    assert str(control["A2"].value).startswith("项目：")
    assert control["G12"].value == "刷新数据包"
    assert control["A21"].value == "主线对照图"
    assert control.freeze_panes == "A11"
    assert len(result["command_palette"]) == 8
    assert result["command_palette"][0]["label"] == "刷新数据包"


def test_excel_export_is_repeatable(synthetic_project) -> None:
    project = synthetic_project["project"]
    repo_root = synthetic_project["paths"].root
    _seed_console_state(project, repo_root=repo_root)

    first = run_excel_export(project, repo_root=repo_root, probe_vba=False)
    second = run_excel_export(project, repo_root=repo_root, probe_vba=False)

    assert Path(first["feed_files"]["overview"]).read_text(encoding="utf-8-sig") == Path(
        second["feed_files"]["overview"]
    ).read_text(encoding="utf-8-sig")
    assert Path(second["workbook_path"]).exists()
